from fastapi import APIRouter, Depends, Request, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import get_db
from app.core.deps import require_roles
from app.models.user import User
from app.schemas.user import (
    CreateUserRequest, UpdateUserRequest, UserResponse, UserListResponse,
    ImportResult,
)
from app.services.user_service import UserService
from app.services.audit_service import AuditService
from app.utils.response import success, error
import openpyxl
import io

router = APIRouter(prefix="/admin/users", tags=["用户管理"], dependencies=[Depends(require_roles(["super_admin"]))])


@router.get("")
async def list_users(
    page: int = 1,
    page_size: int = 20,
    keyword: str = None,
    role: str = None,
    is_active: bool = None,
    db: AsyncSession = Depends(get_db),
):
    """用户列表"""
    service = UserService(db)
    users, total = await service.list_users(page=page, page_size=page_size, keyword=keyword, role=role, is_active=is_active)
    items = [UserResponse.model_validate(u).model_dump(mode='json') for u in users]
    return success(data={"items": items, "total": total, "page": page, "page_size": page_size})


@router.post("")
async def create_user(
    data: CreateUserRequest,
    request: Request,
    current_user: User = Depends(require_roles(["super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    """创建用户"""
    try:
        service = UserService(db)
        user, default_password = await service.create_user(data)
        # 记录日志
        audit = AuditService(db)
        await audit.log(
            operator_id=current_user.user_id,
            operator_name=current_user.name,
            action="user_create",
            resource_type="users",
            resource_id=user.user_id,
            detail={"username": user.username, "role": user.role},
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )
        return success(data={
            "user": UserResponse.model_validate(user).model_dump(mode='json'),
            "default_password": default_password,
        })
    except ValueError as e:
        return error(message=str(e))


@router.put("/{user_id}")
async def update_user(
    user_id: int,
    data: UpdateUserRequest,
    request: Request,
    current_user: User = Depends(require_roles(["super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    """更新用户信息"""
    try:
        service = UserService(db)
        user = await service.update_user(user_id, data)
        audit = AuditService(db)
        await audit.log(
            operator_id=current_user.user_id,
            operator_name=current_user.name,
            action="user_update",
            resource_type="users",
            resource_id=user_id,
            detail=data.model_dump(exclude_unset=True),
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )
        return success(data=UserResponse.model_validate(user).model_dump(mode='json'))
    except ValueError as e:
        return error(message=str(e))


@router.put("/{user_id}/toggle-active")
async def toggle_active(
    user_id: int,
    request: Request,
    current_user: User = Depends(require_roles(["super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    """启用/禁用账号"""
    if user_id == current_user.user_id:
        return error(message="不能禁用自己的账号")
    try:
        service = UserService(db)
        user = await service.toggle_active(user_id)
        audit = AuditService(db)
        await audit.log(
            operator_id=current_user.user_id,
            operator_name=current_user.name,
            action="user_toggle_active",
            resource_type="users",
            resource_id=user_id,
            detail={"is_active": user.is_active},
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )
        return success(data=UserResponse.model_validate(user).model_dump(mode='json'))
    except ValueError as e:
        return error(message=str(e))


@router.post("/import")
async def import_users(
    file: UploadFile = File(...),
    request: Request = None,
    current_user: User = Depends(require_roles(["super_admin"])),
    db: AsyncSession = Depends(get_db),
):
    """批量导入学生"""
    if not file.filename.endswith((".xlsx", ".xls")):
        return error(message="文件格式错误，请上传 .xlsx 或 .xls 文件")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        # 解析表头
        headers = [cell.value for cell in ws[1]]
        required_fields = ["username", "name", "department"]
        for field in required_fields:
            if field not in headers:
                return error(message=f"缺少必填列：{field}")

        # 解析数据
        users_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            if row_data.get("username") and str(row_data["username"]).strip():
                users_data.append({
                    k: str(v).strip() if v else "" for k, v in row_data.items()
                })

        if not users_data:
            return error(message="文件中没有有效数据")

        service = UserService(db)
        result = await service.batch_import(users_data)

        # 记录日志
        audit = AuditService(db)
        await audit.log(
            operator_id=current_user.user_id,
            operator_name=current_user.name,
            action="user_batch_import",
            resource_type="users",
            detail=result,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )

        return success(data=result)
    except Exception as e:
        return error(message=f"导入失败：{str(e)}")


@router.get("/import/template")
async def download_template():
    """下载导入模板"""
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    headers = ["username", "name", "department", "education_level", "grade", "email", "school_id"]
    ws.append(headers)
    ws.append(["2024001", "张三", "计算机学院", "本科", "2024", "zhangsan@xxx.edu.cn", "2024001"])
    ws.append(["2024002", "李四", "计算机学院", "本科", "2024", "lisi@xxx.edu.cn", "2024002"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_import_template.xlsx"},
    )